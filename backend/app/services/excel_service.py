import pandas as pd
import numpy as np
from typing import List, Dict, Any, BinaryIO, Optional, Tuple
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelService:
    # 表头关键词，用于自动检测表头行
    # 注意：关键词按优先级排序，越精确的放前面
    HEADER_KEYWORDS = {
        'product': ['产品名称', '商品名称', '品名'],  # 移除过于宽泛的"名称"和"产品"
        'spec': ['型号规格', '规格型号', '产品描述', '规格', '型号', '描述'],
        'quantity': ['采购数量', '需求数量', '数量'],
        'category': ['小类', '设备分类', '分类', '类别'],
        'unit': ['单位'],
    }
    
    # 需要精确匹配的关键词（完全相等，不是包含关系）
    EXACT_MATCH_KEYWORDS = {'产品名称', '商品名称', '品名', '产品描述', '型号规格', '规格型号'}
    
    def _find_header_row(self, ws) -> Tuple[int, Dict[str, int]]:
        """
        自动检测表头行位置，返回 (行号, 列名映射)
        """
        max_search_rows = min(20, ws.max_row)  # 最多搜索前20行
        best_match = None  # (row_idx, found_columns, column_mapping, match_score)
        
        for row_idx in range(1, max_search_rows + 1):
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    value = str(value).strip()
                row_values.append(value)
            
            # 检查这一行是否包含表头关键词
            found_columns = {}
            match_score = 0
            for col_idx, value in enumerate(row_values):
                if not value:
                    continue
                
                # 检查是否匹配任何关键词
                for field, keywords in self.HEADER_KEYWORDS.items():
                    if field in found_columns:
                        continue  # 已找到该字段，跳过
                    for keyword in keywords:
                        # 精确匹配检查
                        is_exact_match = (value == keyword)
                        # 对于精确匹配关键词，必须完全匹配
                        if keyword in self.EXACT_MATCH_KEYWORDS:
                            if is_exact_match:
                                found_columns[field] = col_idx
                                match_score += 2  # 精确匹配得分更高
                                break
                        else:
                            # 对于其他关键词，允许包含匹配
                            if keyword in value:
                                found_columns[field] = col_idx
                                match_score += 1
                                break
            
            # 如果找到了产品名称列，考虑作为候选
            if 'product' in found_columns:
                # 构建完整的列名映射
                column_mapping = {}
                for col_idx, value in enumerate(row_values):
                    if value:
                        column_mapping[col_idx] = value
                
                # 选择匹配度最高的行
                if best_match is None or match_score > best_match[3]:
                    best_match = (row_idx, found_columns, column_mapping, match_score)
        
        if best_match:
            return best_match[0], best_match[2]
        
        # 未找到表头，默认第一行
        return 1, {}
    
    def _get_merged_cell_value(self, ws, row, col):
        """获取单元格值，处理合并单元格"""
        cell = ws.cell(row=row, column=col)
        
        # 检查是否在合并区域内
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 返回合并区域左上角的值
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        
        return cell.value

    def parse_inventory_excel(self, file: BinaryIO) -> List[Dict[str, Any]]:
        """Parse inventory Excel file"""
        df = pd.read_excel(file)
        
        # Map Chinese column names to English
        column_mapping = {
            '产品名称': 'product_name',
            '设备分类': 'category',
            '分类别名': 'category_alias',
            '型号规格': 'spec',
            '数量': 'quantity',
            '单位': 'unit',
            '销售单价': 'sale_price',
            '销售总价': 'sale_total',
            '合同备注': 'contract_remark',
            '采购单价': 'purchase_price',
            '采购备注': 'purchase_remark',
            '供应商渠道': 'supplier'
        }
        
        # Rename columns
        df = df.rename(columns=column_mapping)
        
        # Replace NaN with None for proper JSON/SQL handling
        df = df.replace({np.nan: None})
        
        # Handle sale_total column (may contain comma-separated values)
        if 'sale_total' in df.columns:
            def parse_sale_total(x):
                if x is None or x == '':
                    return None
                try:
                    return float(str(x).replace(',', '').replace('，', ''))
                except:
                    return None
            df['sale_total'] = df['sale_total'].apply(parse_sale_total)
        
        # Convert to list of dicts
        records = df.to_dict('records')
        
        # Filter out empty rows
        records = [
            r for r in records 
            if r.get('product_name') and str(r['product_name']).strip()
        ]
        
        return records
    
    def parse_procurement_excel(self, file: BinaryIO) -> List[Dict[str, Any]]:
        """
        解析采购需求Excel文件
        支持包含图片、合并单元格、表头不在第一行的情况
        """
        try:
            # 使用 openpyxl 加载工作簿
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            # 自动检测表头行
            header_row, column_names = self._find_header_row(ws)
            
            if not column_names:
                # 如果没找到表头，尝试用 pandas 默认方式
                file.seek(0)
                return self._parse_procurement_excel_fallback(file)
            
            # 构建列索引映射
            col_mapping = {}
            for col_idx, col_name in column_names.items():
                col_name_lower = col_name.lower() if col_name else ''
                
                # 产品名称
                for kw in self.HEADER_KEYWORDS['product']:
                    if kw in col_name:
                        col_mapping['name'] = col_idx
                        break
                
                # 规格描述
                for kw in self.HEADER_KEYWORDS['spec']:
                    if kw in col_name:
                        col_mapping['spec'] = col_idx
                        break
                
                # 数量
                for kw in self.HEADER_KEYWORDS['quantity']:
                    if kw in col_name:
                        col_mapping['quantity'] = col_idx
                        break
                
                # 分类
                for kw in self.HEADER_KEYWORDS['category']:
                    if kw in col_name:
                        col_mapping['category'] = col_idx
                        break
                
                # 单位
                for kw in self.HEADER_KEYWORDS['unit']:
                    if kw in col_name:
                        col_mapping['unit'] = col_idx
                        break
            
            # 读取数据行
            records = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                record = {}
                
                # 获取产品名称
                if 'name' in col_mapping:
                    value = self._get_merged_cell_value(ws, row_idx, col_mapping['name'] + 1)
                    if value is not None:
                        record['name'] = str(value).strip()
                
                # 获取规格
                if 'spec' in col_mapping:
                    value = self._get_merged_cell_value(ws, row_idx, col_mapping['spec'] + 1)
                    if value is not None:
                        record['spec'] = str(value).strip()
                
                # 获取数量
                if 'quantity' in col_mapping:
                    value = self._get_merged_cell_value(ws, row_idx, col_mapping['quantity'] + 1)
                    if value is not None:
                        try:
                            # 处理可能的数字格式
                            if isinstance(value, (int, float)):
                                record['quantity'] = float(value)
                            else:
                                # 尝试解析字符串数字
                                cleaned = str(value).replace(',', '').replace('，', '').strip()
                                if cleaned and cleaned not in ['0', '-', '']:
                                    record['quantity'] = float(cleaned)
                        except (ValueError, TypeError):
                            pass
                
                # 获取分类
                if 'category' in col_mapping:
                    value = self._get_merged_cell_value(ws, row_idx, col_mapping['category'] + 1)
                    if value is not None:
                        record['category'] = str(value).strip()
                
                # 获取单位
                if 'unit' in col_mapping:
                    value = self._get_merged_cell_value(ws, row_idx, col_mapping['unit'] + 1)
                    if value is not None:
                        record['unit'] = str(value).strip()
                
                # 只添加有产品名称的记录
                if record.get('name') and record['name'] not in ['', 'None', 'nan']:
                    records.append(record)
            
            wb.close()
            return records
            
        except Exception as e:
            # 如果 openpyxl 解析失败，回退到 pandas
            print(f"openpyxl 解析失败: {e}，尝试 pandas 方式")
            file.seek(0)
            return self._parse_procurement_excel_fallback(file)
    
    def _parse_procurement_excel_fallback(self, file: BinaryIO) -> List[Dict[str, Any]]:
        """使用 pandas 的回退解析方法"""
        # 尝试多种方式读取
        for header_row in [0, 1, 2, 3, 4, 5, 6, 7]:
            try:
                file.seek(0)
                df = pd.read_excel(file, header=header_row)
                
                # 检查是否找到有效的列
                has_valid_columns = False
                for col in df.columns:
                    col_str = str(col)
                    for keywords in self.HEADER_KEYWORDS.values():
                        for kw in keywords:
                            if kw in col_str:
                                has_valid_columns = True
                                break
                
                if has_valid_columns:
                    break
            except Exception:
                continue
        else:
            # 默认使用第一行作为表头
            file.seek(0)
            df = pd.read_excel(file)
        
        # 解析数据
        records = []
        for _, row in df.iterrows():
            record = {}
            
            # 产品名称
            for col in self.HEADER_KEYWORDS['product']:
                if col in df.columns and pd.notna(row.get(col)):
                    record['name'] = str(row[col]).strip()
                    break
            
            # 规格描述
            for col in self.HEADER_KEYWORDS['spec']:
                if col in df.columns and pd.notna(row.get(col)):
                    record['spec'] = str(row[col]).strip()
                    break
            
            # 数量
            for col in self.HEADER_KEYWORDS['quantity']:
                if col in df.columns and pd.notna(row.get(col)):
                    try:
                        value = row[col]
                        if isinstance(value, (int, float)):
                            record['quantity'] = float(value)
                        else:
                            cleaned = str(value).replace(',', '').replace('，', '').strip()
                            if cleaned:
                                record['quantity'] = float(cleaned)
                    except:
                        pass
                    break
            
            if record.get('name') and record['name'] not in ['', 'None', 'nan']:
                records.append(record)
        
        return records
    
    def generate_inventory_template(self) -> BytesIO:
        """Generate inventory Excel template"""
        columns = [
            '产品名称', '设备分类', '分类别名', '型号规格', '数量', '单位',
            '销售单价', '销售总价', '合同备注', '采购单价', '采购备注', '供应商渠道'
        ]
        df = pd.DataFrame(columns=columns)
        
        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return output


excel_service = ExcelService()
